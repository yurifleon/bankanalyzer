#!/usr/bin/env python3

import csv
import argparse
import re
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit(
        "ERROR: This script requires openpyxl. Install it with:\n"
        "  sudo apt update && sudo apt install python3-openpyxl\n"
        "or inside a venv:\n"
        "  python3 -m venv venv\n"
        "  source venv/bin/activate\n"
        "  pip install openpyxl"
    )


NOISE_WORDS = {
    "POS", "PURCHASE", "DEBIT", "CARD", "CHECKCARD", "VISA", "MASTERCARD",
    "MC", "DISCOVER", "AMEX", "AMERICAN", "EXPRESS", "AUTH", "AUTHORIZATION", "APPROVED", "PENDING",
    "TRANSACTION", "PAYMENT", "PMT", "ACH", "WEB", "ONLINE", "TRANSFER",
    "RECURRING", "RECUR", "AUTOPAY", "AUTO", "BILL", "PAY", "MOBILE",
    "PUR", "PIN", "SIG", "DBT", "CRD", "WITHDRAWAL", "DEPOSIT",
    # Payment processor prefixes that appear before the actual vendor name
    "GGLPAY", "APLPAY", "PYPL", "VENMO", "ZELLE",
}

# Pre-compiled patterns for clean_vendor_name
_RE_SEPARATORS  = re.compile(r"[*#:/\\|,_\-.]+")
_RE_PHONE_SPACE = re.compile(r"\b\d{3}\s*\d{3}\s*\d{4}\b")
_RE_PHONE_DASH  = re.compile(r"\b\d{3}[-.]\d{3}[-.]\d{4}\b")
_RE_DATE_SLASH  = re.compile(r"\b\d{1,2}[/-]\d{1,2}([/-]\d{2,4})?\b")
_RE_DATE_ISO    = re.compile(r"\b\d{4}-\d{2}-\d{2}\b")
_RE_ALPHANUM_ID = re.compile(r"\b[A-Z]*\d+[A-Z0-9]*\b")
_RE_NUMBER      = re.compile(r"\b\d+\b")
_RE_STATE       = re.compile(
    r"\b(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|MD|MA|MI|MN|"
    r"MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|"
    r"WA|WV|WI|WY|DC)\b$"
)
_RE_TOKENS      = re.compile(r"[A-Z]+")

_VENDOR_REPLACEMENTS = (
    ("AMAZON COM", "AMAZON"),
    ("WALMART", "WAL MART"),
    ("WM SUPERCENTER", "WAL MART"),
    ("MCDONALD S", "MCDONALDS"),
    ("BILINTERNET", "INTERNET"),
    ("TST ", ""),
    ("SQ ", ""),
    ("SP ", ""),
)

_KEEP_SHORT = {"BP", "QT", "GE", "HP", "LG", "GM", "KFC", "UPS", "USPS", "IRS"}

_LOCATION_WORDS = {
    "PENSACOLA", "GULF", "BREEZE", "MILTON", "CANTONMENT", "PACE",
    "MOBILE", "SEATTLE", "ATLANTA", "CHICAGO", "NEW", "YORK", "MIAMI",
    "TAMPA", "ORLANDO", "JACKSONVILLE", "NINE", "MILE"
}


def open_csv_with_fallback(input_csv):
    encodings_to_try = [
        "utf-8-sig",
        "utf-16",
        "cp1252",
        "latin-1",
    ]

    for enc in encodings_to_try:
        csvfile = None

        try:
            csvfile = open(input_csv, newline="", encoding=enc)
            csvfile.readline()
            csvfile.seek(0)
            print(f"Using file encoding: {enc}")
            return csvfile

        except UnicodeDecodeError:
            if csvfile:
                csvfile.close()

        except FileNotFoundError:
            raise FileNotFoundError(f"Input file not found: {input_csv}")

        except Exception:
            if csvfile:
                csvfile.close()

    raise RuntimeError("Could not determine CSV encoding.")


def parse_amount(value):
    try:
        value = str(value).replace("$", "").replace(",", "").strip()
        if value == "":
            return Decimal("0")
        if value.startswith("(") and value.endswith(")"):
            value = "-" + value[1:-1]
        return Decimal(value)
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def parse_date(value):
    if not value:
        return None

    value = str(value).strip()

    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"]:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass

    return None


def month_key(tx_date):
    return tx_date.strftime("%Y-%m")


def clean_vendor_name(description):
    text = str(description).upper().strip()
    text = _RE_SEPARATORS.sub(" ", text)

    for old, new in _VENDOR_REPLACEMENTS:
        text = text.replace(old, new)

    text = _RE_PHONE_SPACE.sub(" ", text)
    text = _RE_PHONE_DASH.sub(" ", text)
    text = _RE_DATE_SLASH.sub(" ", text)
    text = _RE_DATE_ISO.sub(" ", text)
    text = _RE_ALPHANUM_ID.sub(" ", text)
    text = _RE_NUMBER.sub(" ", text)
    text = _RE_STATE.sub(" ", text)

    tokens = _RE_TOKENS.findall(text)
    tokens = [t for t in tokens if t not in NOISE_WORDS]
    tokens = [t for t in tokens if len(t) >= 3 or t in _KEEP_SHORT]

    if not tokens:
        return str(description).strip()

    vendor_tokens = tokens[:4]

    while len(vendor_tokens) > 1 and vendor_tokens[-1] in _LOCATION_WORDS:
        vendor_tokens.pop()

    return " ".join(vendor_tokens).strip()


def safe_filename(value):
    safe = "".join(c if c.isalnum() else "_" for c in str(value))
    safe = re.sub(r"_+", "_", safe).strip("_")
    return safe[:60] or "search"


def read_transactions(input_csv, date_col, text_col, debit_col, credit_col, card_col=None):
    transactions = []

    required_cols = [date_col, text_col, debit_col]
    if credit_col is not None and credit_col >= 0:
        required_cols.append(credit_col)
    if card_col is not None and card_col >= 0:
        required_cols.append(card_col)
    required_col = max(required_cols)

    with open_csv_with_fallback(input_csv) as csvfile:
        reader = csv.reader(csvfile)

        for row_number, row in enumerate(reader, start=1):
            if len(row) <= required_col:
                continue

            tx_date = parse_date(row[date_col])

            if not tx_date:
                continue

            description = row[text_col].strip()
            vendor = clean_vendor_name(description)
            amount = parse_amount(row[debit_col])

            if credit_col is not None and credit_col >= 0:
                debit = amount
                credit = parse_amount(row[credit_col])
            else:
                # Credit-card profile:
                # Column E is a single Amount column.
                # Positive amounts are charges/debits.
                # Negative amounts are payments/credits.
                if amount >= 0:
                    debit = amount
                    credit = Decimal("0")
                else:
                    debit = Decimal("0")
                    credit = abs(amount)

            card_number = ""

            if card_col is not None and card_col >= 0:
                try:
                    card_number = row[card_col].strip()
                except Exception:
                    card_number = ""

            transactions.append({
                "row_number": row_number,
                "date": tx_date,
                "month": month_key(tx_date),
                "description": description,
                "vendor": vendor,
                "card_number": card_number,
                "debit": debit,
                "credit": credit,
                "net": credit - debit,
                "row": row,
            })

    return transactions


def get_available_months(transactions):
    return sorted({tx["month"] for tx in transactions})


def choose_month(transactions):
    months = get_available_months(transactions)

    print("\nAvailable months:\n")
    print("0. Whole year / all months")

    for idx, month in enumerate(months, start=1):
        print(f"{idx}. {month}")

    choice = input("\nSelect month number or press Enter for whole year: ").strip()

    if choice == "" or choice == "0":
        return None

    if choice.isdigit():
        idx = int(choice) - 1

        if 0 <= idx < len(months):
            return months[idx]

    print("Invalid month selection. Using whole year.")
    return None


def filter_transactions(transactions, search_text=None, selected_month=None):
    filtered = []
    search_lower = search_text.lower() if search_text else None

    for tx in transactions:
        if selected_month and tx["month"] != selected_month:
            continue

        if search_lower:
            raw_match = search_lower in tx["description"].lower()
            vendor_match = search_lower in tx["vendor"].lower()

            if not raw_match and not vendor_match:
                continue

        filtered.append(tx)

    return filtered


def show_top_patterns(transactions, top_n=30, selected_month=None):
    counter = Counter()
    examples = defaultdict(list)

    for tx in transactions:
        if selected_month and tx["month"] != selected_month:
            continue

        vendor = tx["vendor"]

        if vendor:
            counter[vendor] += 1

            if len(examples[vendor]) < 2 and tx["description"] not in examples[vendor]:
                examples[vendor].append(tx["description"])

    label = selected_month if selected_month else "whole year"
    print(f"\nTop {top_n} likely vendors/business patterns for {label}:\n")

    top_items = counter.most_common(top_n)

    for idx, (pattern, count) in enumerate(top_items, start=1):
        sample = examples[pattern][0] if examples[pattern] else pattern
        print(f"{idx}. {pattern} ({count} rows)")
        print(f"   example: {sample}")

    print()

    return top_items


def show_top_10_by_month(transactions):
    grouped = defaultdict(Counter)

    for tx in transactions:
        grouped[tx["month"]][tx["vendor"]] += 1

    print("\nTop 10 vendors by month:\n")

    for month in sorted(grouped):
        print(month)

        for idx, (vendor, count) in enumerate(grouped[month].most_common(10), start=1):
            print(f"  {idx}. {vendor} ({count} rows)")

        print()


def summarize_by_month_vendor(transactions):
    grouped = {}

    for tx in transactions:
        key = (tx["month"], tx["vendor"], tx.get("card_number", ""))

        if key not in grouped:
            grouped[key] = {
                "month": tx["month"],
                "vendor": tx["vendor"],
                "card_number": tx.get("card_number", ""),
                "count": 0,
                "total_debit": Decimal("0"),
                "total_credit": Decimal("0"),
                "first_date": tx["date"],
                "last_date": tx["date"],
                "examples": [],
            }

        item = grouped[key]
        item["count"] += 1
        item["total_debit"] += tx["debit"]
        item["total_credit"] += tx["credit"]
        item["first_date"] = min(item["first_date"], tx["date"])
        item["last_date"] = max(item["last_date"], tx["date"])

        if len(item["examples"]) < 3 and tx["description"] not in item["examples"]:
            item["examples"].append(tx["description"])

    return sorted(
        grouped.values(),
        key=lambda x: (
            x["month"],
            x["vendor"].lower()
        )
    )


def summarize_month_totals(transactions):
    grouped = {}

    for tx in transactions:
        key = tx["month"]

        if key not in grouped:
            grouped[key] = {
                "month": key,
                "count": 0,
                "total_debit": Decimal("0"),
                "total_credit": Decimal("0"),
                "first_date": tx["date"],
                "last_date": tx["date"],
            }

        item = grouped[key]
        item["count"] += 1
        item["total_debit"] += tx["debit"]
        item["total_credit"] += tx["credit"]
        item["first_date"] = min(item["first_date"], tx["date"])
        item["last_date"] = max(item["last_date"], tx["date"])

    return [grouped[key] for key in sorted(grouped)]


def top_10_per_month(month_vendor):
    by_month = defaultdict(list)

    for item in month_vendor:
        by_month[item["month"]].append(item)

    rows = []

    for month in sorted(by_month):
        ranked = sorted(
            by_month[month],
            key=lambda x: (x["total_debit"], x["count"]),
            reverse=True
        )[:10]

        for rank, item in enumerate(ranked, start=1):
            row = dict(item)
            row["rank"] = rank
            rows.append(row)

    return rows


def autosize_sheet(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def style_header(ws):
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_month_totals_sheet(wb, transactions):
    ws = wb.active
    ws.title = "Monthly Totals"

    ws.append([
        "Month", "Transaction Count", "Total Debit / Charges", "Total Credit / Payments",
        "Net Credit - Debit", "First Date", "Last Date"
    ])

    for item in summarize_month_totals(transactions):
        ws.append([
            item["month"],
            item["count"],
            float(item["total_debit"]),
            float(item["total_credit"]),
            float(item["total_credit"] - item["total_debit"]),
            item["first_date"].isoformat(),
            item["last_date"].isoformat(),
        ])

    style_header(ws)
    autosize_sheet(ws)


def add_monthly_grouped_sheet(wb, month_vendor):
    ws = wb.create_sheet("Monthly Grouped")

    ws.append([
        "Month", "Vendor / Group",
        "Card Number", "Transaction Count", "Total Debit / Charges",
        "Total Credit / Payments", "Net Credit - Debit", "First Date", "Last Date", "Examples"
    ])

    for item in month_vendor:
        ws.append([
            item["month"],
            item["vendor"],
            item.get("card_number", ""),
            item["count"],
            float(item["total_debit"]),
            float(item["total_credit"]),
            float(item["total_credit"] - item["total_debit"]),
            item["first_date"].isoformat(),
            item["last_date"].isoformat(),
            " | ".join(item["examples"]),
        ])

    style_header(ws)
    autosize_sheet(ws)


def add_top_10_sheet(wb, month_vendor):
    ws = wb.create_sheet("Top 10 Per Month")

    ws.append([
        "Month", "Rank", "Vendor / Group",
        "Card Number", "Transaction Count",
        "Total Debit / Charges", "Total Credit / Payments", "Net Credit - Debit", "First Date", "Last Date"
    ])

    for item in top_10_per_month(month_vendor):
        ws.append([
            item["month"],
            item["rank"],
            item["vendor"],
            item["count"],
            float(item["total_debit"]),
            float(item["total_credit"]),
            float(item["total_credit"] - item["total_debit"]),
            item["first_date"].isoformat(),
            item["last_date"].isoformat(),
        ])

    style_header(ws)
    autosize_sheet(ws)


def write_workbook(transactions, output_xlsx):
    wb = Workbook()
    month_vendor = summarize_by_month_vendor(transactions)
    add_month_totals_sheet(wb, transactions)
    add_monthly_grouped_sheet(wb, month_vendor)
    add_top_10_sheet(wb, month_vendor)
    wb.save(output_xlsx)


def write_subset_csv(transactions, output_csv):
    with open(output_csv, "w", newline="", encoding="utf-8") as outfile:
        writer = csv.writer(outfile)

        for tx in transactions:
            writer.writerow(tx["row"])


def print_summary(transactions, title, search_text=None, selected_month=None):
    total_debit = sum((tx["debit"] for tx in transactions), Decimal("0"))
    total_credit = sum((tx["credit"] for tx in transactions), Decimal("0"))
    dates = [tx["date"] for tx in transactions]

    print(f"\n{title}")
    print("-" * len(title))

    if search_text:
        print(f"Search text: {search_text}")

    print(f"Period: {selected_month if selected_month else 'whole year / all months'}")
    print(f"Rows: {len(transactions)}")
    print(f"Total debit/charges/amount: {total_debit}")
    print(f"Total credit/payments: {total_credit}")
    print(f"Net credit minus debit: {total_credit - total_debit}")

    if dates:
        print(f"Transaction timeframe: {min(dates)} to {max(dates)}")
    else:
        print("Transaction timeframe: no valid dates found")


def main():
    parser = argparse.ArgumentParser(
        description=(
            "First generate a full monthly totals workbook, then optionally generate "
            "a separate search-filtered workbook. Supports bank and credit-card CSV layouts."
        )
    )

    parser.add_argument("input_csv", help="Input CSV file")
    parser.add_argument("search_text", nargs="?", default=None, help="Optional text/vendor to search for")

    parser.add_argument("--menu", action="store_true", help="Show top vendor menu for search workbook")
    parser.add_argument("--month-menu", action="store_true", help="Offer whole-year or per-month selection for search workbook")
    parser.add_argument("--month", default=None, help="Search a specific month in YYYY-MM format, for example 2025-03")
    parser.add_argument("--top", type=int, default=30, help="Number of top vendor patterns to show, default 30")
    parser.add_argument("--show-monthly-top10", action="store_true", help="Print top 10 vendors per month")

    parser.add_argument(
        "--profile",
        choices=["bank", "credit-card"],
        default="bank",
        help=(
            "CSV layout profile. "
            "bank = date column B, description 4th column, debit 5th, credit 6th. "
            "credit-card = date column A, description column B, amount column E."
        )
    )

    parser.add_argument(
        "--monthly-output",
        default=None,
        help="Full monthly analysis Excel file. Default: INPUT_monthly_totals.xlsx"
    )

    parser.add_argument(
        "--search-output",
        default=None,
        help="Search-filtered Excel file. Default generated from search/month/menu."
    )

    parser.add_argument("--subset-csv", default=None, help="Optional output CSV for search-filtered rows")

    parser.add_argument("--date-col", type=int, default=None, help="0-based date column index. Overrides profile default.")
    parser.add_argument("--text-col", type=int, default=None, help="0-based description column index. Overrides profile default.")
    parser.add_argument("--debit-col", type=int, default=None, help="0-based debit/amount column index. Overrides profile default.")
    parser.add_argument("--credit-col", type=int, default=None, help="0-based credit column index. Overrides profile default. Use -1 for none.")

    parser.add_argument(
        "--card-col",
        type=int,
        default=None,
        help=(
            "Optional 0-based card/account column index. "
            "Included in outputs and grouping when provided."
        )
    )

    args = parser.parse_args()

    if args.profile == "credit-card":
        default_date_col = 0
        default_text_col = 1
        default_debit_col = 4
        default_credit_col = -1
    else:
        default_date_col = 1
        default_text_col = 3
        default_debit_col = 4
        default_credit_col = 5

    date_col = args.date_col if args.date_col is not None else default_date_col
    text_col = args.text_col if args.text_col is not None else default_text_col
    debit_col = args.debit_col if args.debit_col is not None else default_debit_col
    credit_col = args.credit_col if args.credit_col is not None else default_credit_col

    print(f"Using profile: {args.profile}")
    print(f"Columns: date={date_col}, description={text_col}, debit/amount={debit_col}, credit={credit_col}")

    transactions = read_transactions(
        args.input_csv,
        date_col,
        text_col,
        debit_col,
        credit_col,
        args.card_col
    )

    if not transactions:
        print("ERROR: No valid transactions found.")
        print("Check your column mapping and date format.")
        print("Credit-card profile expects: date=A, description=B, amount=E.")
        return

    input_stem = Path(args.input_csv).stem

    monthly_output = args.monthly_output or f"{input_stem}_monthly_totals.xlsx"

    write_workbook(transactions, monthly_output)
    print_summary(transactions, "Full Monthly Analysis")
    print(f"Full monthly workbook written to: {monthly_output}")

    if args.show_monthly_top10:
        show_top_10_by_month(transactions)

    selected_month = args.month

    if args.month_menu:
        selected_month = choose_month(transactions)

    search_text = args.search_text
    menu_choice_number = None

    if args.menu:
        top_items = show_top_patterns(transactions, top_n=args.top, selected_month=selected_month)

        choice = input(f"Select a number (1-{len(top_items)}) or type custom search text: ").strip()

        if choice.isdigit():
            idx = int(choice) - 1

            if 0 <= idx < len(top_items):
                search_text = top_items[idx][0]
                menu_choice_number = idx + 1
            else:
                print("ERROR: selection out of range.")
                return
        else:
            search_text = choice

    should_make_search_file = bool(search_text or selected_month)

    if not should_make_search_file:
        print("\nNo search/month filter selected. Search workbook was not generated.")
        print("To generate one, run with a search string, --menu, --month, or --month-menu.")
        return

    filtered_transactions = filter_transactions(
        transactions,
        search_text=search_text,
        selected_month=selected_month
    )

    if args.search_output:
        search_output = args.search_output
    else:
        parts = [input_stem, "search"]

        if selected_month:
            parts.append(selected_month)

        if menu_choice_number:
            parts.append(f"option_{menu_choice_number}")

        if search_text:
            parts.append(safe_filename(search_text))

        search_output = "_".join(parts) + ".xlsx"

    write_workbook(filtered_transactions, search_output)
    print_summary(filtered_transactions, "Search / Filtered Analysis", search_text, selected_month)
    print(f"Search workbook written to: {search_output}")

    if args.subset_csv:
        write_subset_csv(filtered_transactions, args.subset_csv)
        print(f"Subset CSV written to: {args.subset_csv}")


if __name__ == "__main__":
    main()
